import logging
import zipfile
import shutil
import os
import re
import pdfplumber
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from flask import Blueprint, request, jsonify, render_template, flash, redirect, url_for, session, send_from_directory
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from modules.database.database import db_blueprint, get_db
from PyPDF2 import PdfMerger

loler_blueprint = Blueprint('loler_blueprint', __name__)

UPLOAD_FOLDER = './input'
OUTPUT_FOLDER = './output'
PROCESSED_FOLDER = './processed'
TEMP_CHUNK_FOLDER = './temp_chunks'
ALLOWED_EXTENSIONS = {'pdf'}
TEMPLATE_PATH = './static/LOLER-ReportTemplate.xlsx'  # Update with the actual path to your template

# Set up logging
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

@loler_blueprint.route('/loler', methods=['GET', 'POST'])
@login_required
def loler():
    success_message = session.pop('success_message', None)
    error_message = session.pop('error_message', None)

    # Fetch the LOLER inspections from the database
    conn = get_db()
    loler_inspections = conn.execute('SELECT * FROM loler_inspections').fetchall()
    view_defects = conn.execute('SELECT * FROM loler_defects').fetchall()
    users = conn.execute('SELECT * FROM users').fetchall()
    defect_count = conn.execute('SELECT COUNT(*) FROM loler_defects').fetchone()[0]
    site_count = conn.execute('SELECT COUNT(*) FROM loler_inspections').fetchone()[0]
    defects = conn.execute('SELECT * FROM loler_defects ORDER BY client_name').fetchall()
    average_processing_time = calculate_average_processing_time(get_db)

    moderate_count, substantial_count, intolerable_count = count_severity_levels(get_db)
    
    conn.close()

    print('Asset Count:', defect_count)
    print('A Count:', moderate_count)
    print('B Count:', substantial_count)
    print('C Count:', intolerable_count)

    return render_template('loler.html', loler_inspections=loler_inspections, users=users, defect_count=defect_count, title='LOLER Reports',
                           success_message=success_message, error_message=error_message, site_count=site_count,
                           moderate_count=moderate_count, substantial_count=substantial_count, intolerable_count=intolerable_count,
                           average_processing_time=average_processing_time, defects=defects)

@loler_blueprint.route('/delete-record-loler', methods=['POST'])
def delete_record_loler():
    loler_report_id = request.form.get('id')
    conn = get_db()
    try:
        # Delete the record from the loler_inspections table
        conn.execute('DELETE FROM loler_inspections WHERE id = ?', (loler_report_id,))
        
        # Also delete associated records from the loler_defects table
        conn.execute('DELETE FROM loler_defects WHERE loler_inspection_id = ?', (loler_report_id,))
        
        conn.commit()
        success_message = "Record successfully deleted."
        error_message = ""
        return jsonify({'success': True, 'message': success_message})
    except Exception as e:
        success_message = ""
        error_message = f"Error deleting Record: {e}"
        return jsonify({'success': False, 'message': error_message}), 500
    finally:
        conn.close()

@loler_blueprint.route('/upload_chunk', methods=['POST'])
def upload_chunk():
    # Example of capturing client name, adjust based on your actual data passing mechanism
    client_name = request.args.get('client_name', 'default_client')
    
    chunk_number = request.form['resumableChunkNumber']
    total_chunks = request.form['resumableTotalChunks']
    file = request.files['file']
    filename = secure_filename(request.form['resumableFilename'])
    chunk_save_path = os.path.join('temp_chunks', f"{filename}.part{chunk_number}")
    file.save(chunk_save_path)

    if all(os.path.exists(os.path.join('temp_chunks', f"{filename}.part{i}")) for i in range(1, int(total_chunks) + 1)):
        # Reassemble the file
        reassembled_file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(reassembled_file_path, 'wb') as outfile:
            for i in range(1, int(total_chunks) + 1):
                chunk_file_path = os.path.join('temp_chunks', f"{filename}.part{i}")
                with open(chunk_file_path, 'rb') as chunk_file:
                    outfile.write(chunk_file.read())
                os.remove(chunk_file_path)  # Clean up after reassembly

        # Assuming your processing function can handle directly the reassembled file path
        # and you adjust it to accept and handle the client_name
        try:
            return jsonify({'status': 'success', 'message': 'File uploaded'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)})
    else:
        return jsonify({'status': 'in_progress', 'message': 'Chunk received'})

@loler_blueprint.route('/start_processing', methods=['POST'])
def start_processing():
    data = request.get_json()
    client_name = data.get('client_name')
    job_number = data.get('job_number')
    client_address_1 = data.get('client_address_1', '')
    client_address_2 = data.get('client_address_2', '')
    client_postcode = data.get('client_postcode', '')
    client_contact = data.get('client_contact', '')
    client_contact_no = data.get('client_contact_no', '')

    if not client_name or not job_number:
        return jsonify({'status': 'error', 'message': 'Client name and Job number are required'}), 400

    try:
        logger.debug("Starting processing for client: %s", client_name)
        xlsx_file_path = process_loler_pdfs(UPLOAD_FOLDER, OUTPUT_FOLDER, client_name, client_address_1, client_address_2, client_postcode, job_number, client_contact_no, client_contact, get_db)
        filename = os.path.basename(xlsx_file_path)

        clear_input_folder(UPLOAD_FOLDER)
        clear_input_folder(TEMP_CHUNK_FOLDER)
        
        download_url = url_for('loler_blueprint.download_file', filename=filename, _external=True)
        
        return jsonify({'status': 'success', 'download_url': download_url})
    except Exception as e:
        logger.error("Error during processing: %s", str(e), exc_info=True)
        clear_input_folder(UPLOAD_FOLDER)
        return jsonify({'status': 'error', 'message': str(e)}), 500

@loler_blueprint.route('/download/<filename>', methods=['GET'])
@login_required
def download_file(filename):
    try:
        return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)
    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({"error": "File not found or another error occurred"}), 404

# Functions from pdfextract.py
def merge_pdfs(pdf_files, output_path):
    merger = PdfMerger()
    for pdf in pdf_files:
        with open(pdf, "rb") as pdf_file:
            merger.append(pdf_file)
    with open(output_path, "wb") as merged_pdf:
        merger.write(merged_pdf)

def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text(x_tolerance=2)
    return text

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def clear_input_folder(upload_folder):
    pdf_files = [file for file in os.listdir(upload_folder) if file.endswith('.pdf')]
    for pdf_file in pdf_files:
        file_path = os.path.join(upload_folder, pdf_file)
        os.remove(file_path)
        print(f"Deleted: {file_path}")

# Generate a unique filename (from lolerextract.py)
def generate_unique_filename(client_name):
    timestamp = datetime.now().strftime("%Y%m%d-%H")
    filename = f"{client_name}_{timestamp}.xlsx"
    return filename

def process_loler_pdfs(input_folder, output_folder, client_name, client_address_1, client_address_2, client_postcode, job_number, client_contact_no, client_contact, get_db):
    try:
        # Create the output directory if it doesn't exist
        os.makedirs(output_folder, exist_ok=True)

        # Variables to track report dates and count
        earliest_next_inspection_date = "N/A"
        report_date = None
        report_count = 0
        skipped_documents = []

        # Generate a unique filename based on client_name
        unique_filename = generate_unique_filename(client_name)
        xlsx_output_file = os.path.join(output_folder, unique_filename)

        # Load the Excel template
        workbook = load_workbook(TEMPLATE_PATH)
        sheet = workbook.active

        # Fill in the client details in the Excel template
        sheet['A7'] = client_name
        sheet['A8'] = client_address_1
        sheet['A9'] = client_address_2
        sheet['A10'] = client_postcode
        sheet['D10'] = job_number
        sheet['D11'] = client_contact_no
        sheet['D12'] = client_contact

        # Write headers to the Excel sheet
        headers = ['Equipment Type', 'ISI Number', 'Serial Number', 'DOM', 'SWL',
                   'Client Ref', 'Location', 'Report ID', 'A Defect', 'B Defect', 'C Defect', 'D Defect',
                   'Report Date', 'Next Inspection Date']
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=14, column=col_num, value=header)

        start_row = 15  # Starting row for data input

        logged_user = f"{current_user.first_name} {current_user.second_name}"

        for filename in os.listdir(input_folder):
            if filename.endswith(".pdf"):
                report_count += 1
                pdf_path = os.path.join(input_folder, filename)
                text = extract_text_from_pdf(pdf_path)

                # Extract information between "### Metadata ###" and "### End Metadata ###"
                metadata_match = re.search(r"### Metadata ###(.*?)### End Metadata ###", text, re.DOTALL)
                if metadata_match:
                    metadata_section = metadata_match.group(1).strip()

                    # Extract relevant information from the metadata section
                    metadata_dict = {}

                    # Define the keys to extract
                    keys_to_extract = ['Equipment Type', 'ISI Number', 'Serial Number', 'DOM', 'SWL',
                                       'Client Ref', 'Location', 'Report ID', 'A Defect', 'B Defect', 'C Defect', 'D Defect']

                    for key in keys_to_extract:
                        key_match = re.search(rf"{key}\s*([^\n]*)", metadata_section)
                        metadata_dict[key] = key_match.group(1).strip() if key_match else ""

                    # Extract report date and next inspection date
                    report_date_match = re.search(r"Report Date\s*([^\n]*)", text)
                    report_date = report_date_match.group(1).strip() if report_date_match else None

                    next_inspection_date_match = re.search(r"Next Inspection Date\s*([^\n]*)", text)
                    next_inspection_date = next_inspection_date_match.group(1).strip() if next_inspection_date_match else None

                    if next_inspection_date is None or next_inspection_date == '' or next_inspection_date == 'Surveyor':
                        # Skip this document
                        skipped_documents.append((filename, next_inspection_date))
                        continue

                    # Write the row to the Excel sheet starting from row 15
                    xlsx_row = [metadata_dict.get(key, "") for key in keys_to_extract] + [report_date, next_inspection_date]
                    for col_num, value in enumerate(xlsx_row, start=1):
                        sheet.cell(row=start_row, column=col_num, value=value)
                    start_row += 1

                    # Convert report_date and next_inspection_date to datetime objects if they are not None
                    if report_date:
                        report_date = datetime.strptime(report_date, '%d/%m/%Y')
                    if next_inspection_date:
                        next_inspection_date = datetime.strptime(next_inspection_date, '%d/%m/%Y')

                    # Update earliest_next_inspection_date if necessary
                    if earliest_next_inspection_date == "N/A" or (next_inspection_date and next_inspection_date < earliest_next_inspection_date):
                        earliest_next_inspection_date = next_inspection_date or earliest_next_inspection_date

                    # Insert data into the "loler_defects" table
                    db_insert_loler_defects(client_name, metadata_dict, report_date, next_inspection_date, get_db, logged_user)

        # Define the table range
        table_end_row = start_row - 1
        table_ref = f"A14:N{table_end_row}"

        # Create the table
        table = Table(displayName="LOLER_Table", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        # Save the populated Excel file
        workbook.save(xlsx_output_file)

        # Insert data into the "loler_inspections" table
        db_insert_loler_inspection(client_name, report_date, earliest_next_inspection_date, report_count, unique_filename, get_db, logged_user)

        return xlsx_output_file

    except Exception as e:
        logger.error("Error processing PDFs: %s", str(e), exc_info=True)
        raise


def db_insert_loler_defects(client_name, metadata_dict, report_date, next_inspection_date, get_db, logged_user):
    try:
        conn = get_db()

        # Convert report_date and next_inspection_date to datetime objects if they are not None
        if report_date and not isinstance(report_date, datetime):
            report_date = datetime.strptime(report_date, '%d/%m/%Y')
        if next_inspection_date and not isinstance(next_inspection_date, datetime):
            next_inspection_date = datetime.strptime(next_inspection_date, '%d/%m/%Y')

        # Convert report_date and next_inspection_date to strings if they are datetime objects
        report_date_str = report_date.strftime('%d-%m-%Y') if isinstance(report_date, datetime) else report_date
        next_inspection_date_str = next_inspection_date.strftime('%d-%m-%Y') if isinstance(next_inspection_date, datetime) else next_inspection_date

        # Capture current time
        process_date = datetime.now().strftime('%d-%m-%Y')

        # Append inspection URL
        inspection_id = f"https://portal.servicesight.com/industrialsafetyinspectionsltd/inspections/{metadata_dict.get('Report ID', '')}"

        # Insert data into the table using correct column names
        conn.execute('''
            INSERT INTO loler_defects (client_name, equipment_type, isi_number, serial_number, date_of_manufacture, safe_working_load, client_id, sub_location, report_id, a_defect, b_defect, c_defect, d_defect, report_date, next_inspection_date, process_date, logged_by) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (client_name, metadata_dict.get('Equipment Type', ''), metadata_dict.get('ISI Number', ''), metadata_dict.get('Serial Number', ''), metadata_dict.get('date_of_manufacture', ''), metadata_dict.get('SWL', ''), metadata_dict.get('Client Ref', ''), metadata_dict.get('Location', ''), inspection_id, metadata_dict.get('A Defect', ''), metadata_dict.get('B Defect', ''), metadata_dict.get('C Defect', ''), metadata_dict.get('D Defect', ''), report_date_str, next_inspection_date_str, process_date, logged_user))
        
        conn.commit()
        logger.debug("Data inserted into loler_defects")
    except Exception as e:
        logger.error("Database operation error: %s", str(e), exc_info=True)
    finally:
        conn.close()

def db_insert_loler_inspection(client_name, report_date, next_inspection_date, report_count, unique_filename, get_db, logged_user):
    try:
        conn = get_db()

        # Convert report_date and next_inspection_date to strings if they are datetime objects
        report_date_str = report_date.strftime('%d/%m/%Y') if isinstance(report_date, datetime) else report_date
        next_inspection_date_str = next_inspection_date.strftime('%d/%m/%Y') if isinstance(next_inspection_date, datetime) else next_inspection_date
        file_name = unique_filename  # Assign unique_filename to file_name

        # Capture current time
        process_date = datetime.now().strftime('%d/%m/%Y')

        # Insert data into the table
        conn.execute('''
            INSERT INTO loler_inspections (client_name, report_date, next_inspection_date, file_name, report_count, logged_by, process_date) 
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (client_name, report_date_str, next_inspection_date_str, file_name, report_count, logged_user, process_date))
        
        conn.commit()
        logger.debug("Data inserted into loler_inspections")
    except Exception as e:
        logger.error("Database operation error: %s", str(e), exc_info=True)
    finally:
        conn.close()

def calculate_average_processing_time(get_db):
    conn = get_db()
    cursor = conn.cursor()

    try:
        # Query the database to fetch inspection_date and process_date
        cursor.execute("SELECT report_date, process_date FROM loler_inspections")
        rows = cursor.fetchall()

        total_time = timedelta(0)
        total_reports = len(rows)

        for row in rows:
            inspection_date = datetime.strptime(row['report_date'], "%d/%m/%Y")
            process_date = datetime.strptime(row['process_date'], "%d/%m/%Y")
            time_difference = process_date - inspection_date
            total_time += time_difference

        # Calculate the average processing time
        if total_reports > 0:
            average_time = total_time / total_reports
        else:
            average_time = timedelta(0)

        return average_time
    except Exception as e:
        logger.error("Error calculating average processing time: %s", str(e), exc_info=True)
        return None
    finally:
        conn.close()

def count_severity_levels(get_db):
    # Connect to the database
    conn = get_db()
    cursor = conn.cursor()

    try:
        # Count occurrences of each severity level
        cursor.execute("SELECT COUNT(*) FROM loler_defects WHERE c_defect != 'None'")
        moderate_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM loler_defects WHERE b_defect != 'None'")
        substantial_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM loler_defects WHERE a_defect != 'None'")
        intolerable_count = cursor.fetchone()[0]

        return moderate_count, substantial_count, intolerable_count
    finally:
        conn.close()
